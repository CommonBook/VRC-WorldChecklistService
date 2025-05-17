'''import xlwt
from xlwt import Workbook'''
import time
import datetime
import dateutil
from dateutil.tz import tzutc
from datetime import tzinfo
import xlsxwriter
from xlsxwriter import Workbook
import openpyxl

class worldID_Data:
    def __init__(self):
        self.data = {}#{'wrld_9d145784-bc19-4237-acf8-984a34b0e4a3': {'author_id': 'usr_26be101b-fa4f-4b26-90f2-7e861e4a48f4', 'author_name': 'Maryo_', 'capacity': 32, 'recommended_capacity': 16, 'created_at': datetime.datetime(2025, 3, 1, 1, 13, 30, 965000, tzinfo=tzutc()), 'description': 'Sad 3D Land․ this is the the original map used for the sad Mario gif Thanks for 1K visitsǃ', 'favorites': 182, 'featured': False, 'heat': 3, 'id': 'wrld_9d145784-bc19-4237-acf8-984a34b0e4a3', 'image_url': 'https://api.vrchat.cloud/api/1/file/file_58c31c82-aad4-4dc2-b38b-89699d34ec20/1/file', 'instances': [['20563~private(usr_a4339c5d-e178-4204-b740-3cbac3ce5596)~region(us)', 1]], 'labs_publication_date': '2025-03-01T05:24:39.552Z', 'name': 'Sad 3D Land', 'namespace': None, 'occupants': 1, 'organization': 'vrchat', 'popularity': 4, 'preview_youtube_id': None, 'private_occupants': 1, 'public_occupants': 0, 'publication_date': '2025-04-08T11:00:14.071Z', 'release_status': 'public', 'tags': ['author_tag_Funny', 'author_tag_Mario', 'author_tag_Meme', 'author_tag_Joke', 'system_approved'], 'thumbnail_image_url': 'https://api.vrchat.cloud/api/1/image/file_58c31c82-aad4-4dc2-b38b-89699d34ec20/1/256', 'unity_packages': [{'id': 'unp_59d0f4e8-165a-4e58-8ec9-472a4d00dbd7', 'asset_url': 'https://api.vrchat.cloud/api/1/file/file_8124aec6-b2cf-4ce7-a880-8aa564144ac3/7/file', 'asset_url_object': None, 'asset_version': 4, 'created_at': datetime.datetime(2025, 4, 14, 9, 39, 29, 539000, tzinfo=tzutc()), 'impostorizer_version': None, 'performance_rating': None, 'platform': 'android', 'plugin_url': None, 'plugin_url_object': None, 'unity_sort_number': 20220322000, 'unity_version': '2022.3.22f1', 'world_signature': 'AH4z96+3qlNPYMRcb5tN1X4MfVkQnVA0FwGq+PoOJbe1PpOjpg==', 'impostor_url': None, 'scan_status': None, 'variant': None}, {'id': 'unp_94333126-b888-43d8-ad5e-2f8ff0e5c345', 'asset_url': 'https://api.vrchat.cloud/api/1/file/file_d3fd8202-9fce-48b8-8f23-ab13a741729a/8/file', 'asset_url_object': None, 'asset_version': 4, 'created_at': datetime.datetime(2025, 4, 14, 9, 37, 22, 147000, tzinfo=tzutc()), 'impostorizer_version': None, 'performance_rating': None, 'platform': 'standalonewindows', 'plugin_url': None, 'plugin_url_object': None, 'unity_sort_number': 20220322000, 'unity_version': '2022.3.22f1', 'world_signature': 'AEA+vd/suAU4wq7AqcpsgRhHidgQuMxyoJlcch7TtlWxwx4NPA==', 'impostor_url': None, 'scan_status': None, 'variant': None}], 'updated_at': datetime.datetime(2025, 4, 14, 9, 39, 29, 754000, tzinfo=tzutc()), 'url_list': [], 'version': 19, 'visits': 2101, 'udon_products': [], 'visited': True}}

    def export(self):
        # This haphazard array lists every datapoint that is not to be exported to the spreadsheet. Basically, it's garbage.
        # If any of this is important to you, then you'll have to save it yourself, since anything not exported is also not loaded later.
        # 'visited' is skipped here because it is a special case.
        skippedPoints = ["visited", "namespace", "occupants", "organization", "preview_youtube_id", "private_occupants", 
                         "id", "defaultContentSettings", "recommended_capacity", "instances", "tags", "unity_packages", "url_list", "udon_products", "assetUrlObject",
                         "pluginUrlObject", "labs_publication_date", "public_occupants"]

        x_index : int = 0
        y_index : int = 0

        workbook = Workbook("data_export.xlsx", {'remove_timezone': True})
        worksheet = workbook.add_worksheet()

        # Formatting junk
        titleFormat = workbook.add_format()
        titleFormat.set_bold()
        titleFormat.set_bg_color('green')
        titleFormat.set_color("white")
        titleFormat.set_font_size(14)
        titleFormat.set_align("center")

        nameFormat = workbook.add_format()
        nameFormat.set_bg_color('yellow')
        nameFormat.set_bold()
        nameFormat.set_align("center")
        nameFormat.set_border_color('gray')
        nameFormat.set_border()

        checkFormat = workbook.add_format()
        checkFormat.set_bg_color('gray')
        checkFormat.set_checkbox(True)
        checkFormat.set_font_color('white')
        checkFormat.set_border_color('white')
        checkFormat.set_border()

        generalFormat = workbook.add_format()
        generalFormat.set_align('center')

        worksheet.set_column(0, 27, 17, generalFormat)
        worksheet.set_column(1, 26, 17)
        worksheet.set_column(0, 0, 9, checkFormat)

        worksheet.write(y_index, x_index, "Visited", titleFormat)
        x_index += 1
        worksheet.write(y_index, x_index, "World ID", titleFormat)
        x_index += 1


        allKeys : list[str] = []

        # No idea how to just get the first key's name as a fallback so we're doing allat.
        names = self.data.keys()
        first = ""
        for value in names:
            first = value
            break

        # Just setting up titles for every data point we output
        dictionary = self.data.get(first)

        keys = dictionary.keys()
        
        for key in keys:
            if key in skippedPoints:
                continue
            
            allKeys.append(key)
            
            text = key.capitalize()
            text = text.replace("_", " ")

            worksheet.write(y_index, x_index, text, titleFormat)
            x_index += 1

        x_index = 0
        y_index = 1

        # Actually writing to the workbook
        for worldID in self.data.keys():
            visited = self.data[worldID]["visited"]
            
            worksheet.insert_checkbox(y_index, x_index, visited, checkFormat)
            #worksheet.write(y_index, x_index, visited)

            x_index += 1
            worksheet.write(y_index, x_index, worldID)
            x_index += 1

            for dataPointType in allKeys:
                if dataPointType in skippedPoints:
                    continue
                
                # print(">> " + str(dataPointType))
                worksheet.write(y_index, x_index, self.data[worldID][dataPointType])

                x_index += 1
            y_index += 1
            x_index = 0

        # Apply post column formatting (not sure if this is actually necessary)
        worksheet.set_column(11, 11, 20, nameFormat)
        worksheet.set_column(0, 0, 9, checkFormat)

        workbook.close()
        print("Successfully exported data.")

    def _add_world(worldID : str, data : dict):
        pass

    def load(self):
        userInput = input("Name of excel document to load (default data_export): ")
        fileName = userInput if userInput != "" else "data_export.xlsx"

        if fileName[-5:] != ".xlsx":
            fileName = fileName + ".xlsx"

        try:
            file = openpyxl.load_workbook(fileName)
            sheet = file.active
        except FileNotFoundError:
            print("File " + fileName + " does not exist.")
            return
        except KeyboardInterrupt:
            print("Stopping import.")
            return

        # Scan every row for each world
        for rowNum in range(1, sheet.max_row):
            dataIn = {}

            # Scan every column to load all of the collected datapoints.
            for column in sheet.iter_cols(1,sheet.max_column):
                dataIn.update({column[0].value.lower().replace(" ", "_"): column[rowNum].value})
            #print(dataIn)
            ID = dataIn.get("world_id")
            dataIn.pop("world_id")

            self.data.update({ID : dataIn})
            #print(self.data)


if __name__ == "__main__":
    new = worldID_Data()
    new.load()
    new.export()
